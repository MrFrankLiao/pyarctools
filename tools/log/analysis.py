import xml.etree.ElementTree as ET
import openpyxl
import itertools
import os
from datetime import datetime


def analysis(start, end, path):
    msg_list = []
    set_list = []
    start = datetime.strptime(start.get(), '%Y-%m-%d')
    end = datetime.strptime(end.get(), '%Y-%m-%d')

    if not path:
        path = './data/arcgis_log/'

    wb = openpyxl.Workbook(write_only=True)
    sh_counts_times = wb.create_sheet(title="total count")
    sh_all = wb.create_sheet(title="total table")
    sh_user_usage_count_2 = wb.create_sheet(title="user total count")
    sh_counts_times.append(["source", "count"])
    sh_all.append(["time", 'type', 'code', "source", "process", "thread",
                   "methodName", "machine", "user", "elapsed", "msg"])
    sh_user_usage_count_2.append(["user", "source", "count"])
    print("start to analysis")
    for fn in os.listdir(path):
        with open('%s/%s' % (path, fn), 'r', encoding='utf-8') as f:
            it = itertools.chain('<root>', f, '</root>')
            datas = ET.fromstringlist(it)
            for data in datas:
                split_time = datetime.strptime(data.attrib['time'].split('T')[0],
                                               '%Y-%m-%d')
                if start <= split_time <= end:
                    if "Request user" in data.text:
                        msg_list.append(data.text)

                    sh_all.append([data.attrib['time'],
                                   data.attrib['type'],
                                   data.attrib['code'],
                                   data.attrib['source'],
                                   data.attrib['process'],
                                   data.attrib['thread'],
                                   data.attrib['methodName'],
                                   data.attrib['machine'],
                                   data.attrib['user'],
                                   data.attrib['elapsed'],
                                   data.text])
    source_list_of_msg_list = []
    for m in msg_list:
        sp1 = m.split(',')
        mapserver = sp1[1].split(':')[1][1:]
        source_list_of_msg_list.append(mapserver)
    from collections import Counter
    counter = Counter(source_list_of_msg_list)
    for i in counter:
        sh_counts_times.append([i, counter[i]])
    for m in msg_list:
        sp1 = m.split(',')
        user = sp1[0].split(':')[1][1:]
        mapserver = sp1[1].split(':')[1][1:]
        set_list.append(user + "$$$" + mapserver)

    from collections import Counter
    counter = Counter(set_list)
    for i in counter:
        a = i.split('$$$')
        sh_user_usage_count_2.append([a[0], a[1], counter[i]])

    wb.save('analysis.xlsx')



